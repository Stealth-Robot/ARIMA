"""Import data from 'lettuce billy joel.xlsx' into the ARIMA database.

Usage:
    flask import-rock

Requires: flask seed and the initial kpop import to have been run first.
The Excel file must be at the project root as 'lettuce billy joel.xlsx'.
"""

import re
import sys

import openpyxl

from app.extensions import db
from app.models.user import User
from app.models.music import (
    Artist, Album, Song, Rating, ArtistSong, AlbumSong, ArtistArtist, album_genres,
)
from app.models.changelog import Changelog
from app.services.artist import generate_unique_slug

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

EXCEL_PATH = 'lettuce billy joel.xlsx'

USER_COLUMNS = {
    1: 'Assy', 2: 'Stealth', 3: 'Deren', 4: 'Diam', 5: 'Emily',
    6: 'Globe', 7: 'Kanjo', 8: 'Quiet', 9: 'Rose', 10: 'Sam', 11: 'Toki',
}

SKIPPED_TABS = {
    'The Beatles', 'Queen', 'Green Day', 'Elvis Presley',
    'Ed Sheeran', 'Billy Joel', 'ABBA',
}

META_TABS = {'Changelog', 'RULES', 'TEMPLATE', 'STATS', 'STATS 2.0'}

TAB_NAME_MAP = {
    'NSYNC': '*NSYNC',
}

# Gender: 0=Female, 1=Male, 2=Mixed
GENDER_MAP = {
    'Adele': 0, 'Ariana Grande': 0, 'Avril Lavigne': 0, 'Billie Eilish': 0,
    'Britney Spears': 0, 'Carly Rae Jepsen': 0, 'Chappell Roan': 0,
    'Demi Lovato': 0, 'Doja Cat': 0, 'Dua Lipa': 0, 'Evanescence': 2,
    'Hayley Williams': 0, 'Lady Gaga': 0, 'Lana Del Rey': 0, 'Leah Kate': 0,
    'Lexie Liu': 0, 'LØLØ': 0, 'Maggie Lindemann': 0, 'MARINA': 0,
    'Mitski': 0, 'Nicki Minaj': 0, 'Olivia Rodrigo': 0, 'Rihanna': 0,
    'Sabrina Carpenter': 0, 'Selena Gomez': 0, 'Taylor Swift': 0,
    '*NSYNC': 1, 'Bruno Mars': 1, 'Lil Nas X': 1, 'Michael Jackson': 1,
    'The Weeknd': 1,
    'Against The Current': 2, 'All Time Low': 1, 'Big Time Rush': 1,
    'Fall Out Boy': 1, 'HOYO-MIX': 2, 'Imagine Dragons': 1,
    'Jonas Brothers': 1, 'Linkin Park': 2, 'LoL': 2,
    'Muse': 1, 'My Chemical Romance': 1, 'Nirvana': 1,
    'One Direction': 1, 'Panic! At The Disco': 1, 'Paramore': 2,
    'Radiohead': 1,
}

ALBUM_YEAR_RE = re.compile(r'\((\d{4})(?:-\d{4})?\)\s*$')
IMPORTRANGE_TAB_RE = re.compile(r'""([^"!]+)!')
IFERROR_FALLBACK_RE = re.compile(r',\s*([^,)]+)\)\s*$')

KNOWN_SUBUNITS = {
    'HOYO-MIX': {'Honkai Impact 3rd', 'Genshin Impact', 'Tears of Themis'},
    'Lana Del Rey': {'REMIXES'},
}


# ---------------------------------------------------------------------------
# Parsing helpers (same as jpop import)
# ---------------------------------------------------------------------------

def _parse_album_year(text):
    m = ALBUM_YEAR_RE.search(text)
    if not m:
        return None
    return text[:m.start()].strip(), m.group(1)


def _is_formula(cell):
    return isinstance(cell.value, str) and cell.value.startswith('=')


def _extract_formula_tab(cell):
    if not _is_formula(cell):
        return None
    m = IMPORTRANGE_TAB_RE.search(cell.value)
    return m.group(1) if m else None


def _extract_formula_fallback(cell):
    if not _is_formula(cell):
        return None
    m = IFERROR_FALLBACK_RE.search(cell.value)
    if not m:
        return None
    raw = m.group(1).strip().strip('"')
    if raw == '' or raw == '""':
        return None
    try:
        return float(raw)
    except ValueError:
        return None


def _parse_rating_cell(cell):
    if _is_formula(cell):
        tab = _extract_formula_tab(cell)
        val = _extract_formula_fallback(cell)
        return val, True, tab
    if isinstance(cell.value, (int, float)):
        v = int(cell.value)
        if 0 <= v <= 5:
            return v, False, None
    return None, False, None


def _get_comment_text(cell):
    if cell.comment:
        return cell.comment.text.strip() or None
    return None


def _extract_song_data(row):
    ratings = {}
    notes = {}
    is_formula_row = False
    primary_tab = None
    for col_idx, username in USER_COLUMNS.items():
        cell = row[col_idx]
        val, is_f, tab = _parse_rating_cell(cell)
        if is_f:
            is_formula_row = True
            if tab:
                primary_tab = tab
        if val is not None:
            ratings[username] = int(val)
        comment = _get_comment_text(cell)
        if comment:
            notes[username] = comment
    return {
        'ratings': ratings, 'notes': notes,
        'is_formula_row': is_formula_row, 'primary_tab': primary_tab,
    }


def _row_has_ratings(row):
    for col_idx in USER_COLUMNS:
        cell = row[col_idx]
        if isinstance(cell.value, (int, float)) and 0 <= cell.value <= 5:
            return True
        if _is_formula(cell):
            return True
    return False


def _parse_tab(ws, known_subunits=None):
    subunits = known_subunits or set()
    in_subunit_tab = bool(subunits)
    items = []
    current_album = None

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        a_cell = row[0]
        q_cell = row[16] if len(row) > 16 else None
        a_val = str(a_cell.value).strip() if a_cell.value else ''
        is_song = q_cell and q_cell.value is True

        if not a_val:
            continue

        if is_song:
            song_data = _extract_song_data(row)
            song_data['name'] = a_val
            if current_album:
                current_album['songs'].append(song_data)
            else:
                current_album = {
                    'type': 'album', 'name': 'Unknown Album', 'year': None,
                    'songs': [song_data], '_implicit': True,
                }
                items.append(current_album)
            continue

        parsed = _parse_album_year(a_val)
        if parsed:
            name, year = parsed
            current_album = {'type': 'album', 'name': name, 'year': year, 'songs': []}
            items.append(current_album)
            continue

        if a_val in subunits:
            items.append({'type': 'subunit', 'name': a_val})
            current_album = None
            continue

        if _row_has_ratings(row):
            song_data = _extract_song_data(row)
            song_data['name'] = a_val
            if current_album:
                current_album['songs'].append(song_data)
            else:
                current_album = {
                    'type': 'album', 'name': 'Unknown Album', 'year': None,
                    'songs': [song_data], '_implicit': True,
                }
                items.append(current_album)
            continue

        if in_subunit_tab:
            song_data = _extract_song_data(row)
            song_data['name'] = a_val
            if current_album:
                current_album['songs'].append(song_data)
            else:
                current_album = {
                    'type': 'album', 'name': 'Unknown Album', 'year': None,
                    'songs': [song_data], '_implicit': True,
                }
                items.append(current_album)

    return items


# ---------------------------------------------------------------------------
# Import logic
# ---------------------------------------------------------------------------

def import_rock():
    """Main entry point."""
    print(f'Loading {EXCEL_PATH} (formulas mode)...')
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=False)
    print(f'  {len(wb.sheetnames)} sheets found')

    user_map = _build_user_map()
    used_slugs = {a.slug for a in Artist.query.all() if a.slug}

    artist_map = {}
    song_lookup = {}
    total_songs = 0
    total_ratings = 0

    all_tabs = set(wb.sheetnames)
    explicit_tabs = sorted(all_tabs - META_TABS - SKIPPED_TABS - {'Misc Artists'})

    for tab_name in explicit_tabs:
        artist_name = TAB_NAME_MAP.get(tab_name, tab_name)
        ws = wb[tab_name]
        items = _parse_tab(ws, known_subunits=KNOWN_SUBUNITS.get(tab_name))

        songs, ratings = _import_artist_tab(
            artist_name, tab_name, items, user_map, artist_map, song_lookup, used_slugs,
        )
        total_songs += songs
        total_ratings += ratings
        print(f'  {tab_name}: {songs} songs, {ratings} ratings')

    linked, formula_created = _link_formula_songs(wb, user_map, artist_map, song_lookup)
    total_songs += formula_created
    print(f'  Multi-artist links: {linked}, formula fallbacks: {formula_created}')

    misc_songs, misc_ratings = _import_misc_artists(wb, user_map, artist_map, song_lookup, used_slugs)
    total_songs += misc_songs
    total_ratings += misc_ratings

    cl_count = _import_changelog(wb)

    print(f'\n=== ROCK IMPORT COMPLETE ===')
    print(f'  Artists: {len(artist_map)}')
    print(f'  Songs: {total_songs}')
    print(f'  Ratings: {total_ratings}')
    print(f'  Multi-artist links: {linked} (+ {formula_created} fallback)')
    print(f'  Changelog: {cl_count}')


def _build_user_map():
    user_map = {}
    for user in User.query.all():
        user_map[user.username] = user.id
    print(f'  {len(user_map)} users found in DB')
    return user_map


def _import_artist_tab(artist_name, tab_name, items, user_map, artist_map, song_lookup, used_slugs):
    gender_id = GENDER_MAP.get(artist_name, 2)

    main_artist = Artist.query.filter_by(name=artist_name).first()
    if not main_artist:
        slug = generate_unique_slug(artist_name, used_slugs)
        used_slugs.add(slug)
        main_artist = Artist(
            name=artist_name, slug=slug, gender_id=gender_id,
            country_id=3, submitted_by_id=0,
        )
        db.session.add(main_artist)
        db.session.flush()

    artist_map[tab_name] = main_artist.id
    artist_map[artist_name] = main_artist.id

    total_songs = 0
    total_ratings = 0
    current_artist_id = main_artist.id

    for item in items:
        if item['type'] == 'subunit':
            sub_name = item['name']
            sub_artist = Artist.query.filter_by(name=sub_name).first()
            if not sub_artist:
                slug = generate_unique_slug(sub_name, used_slugs)
                used_slugs.add(slug)
                sub_artist = Artist(
                    name=sub_name, slug=slug, gender_id=gender_id,
                    country_id=3, submitted_by_id=0,
                )
                db.session.add(sub_artist)
                db.session.flush()
                existing_rel = ArtistArtist.query.filter_by(
                    artist_1=main_artist.id, artist_2=sub_artist.id,
                ).first()
                if not existing_rel:
                    db.session.add(ArtistArtist(
                        artist_1=main_artist.id, artist_2=sub_artist.id, relationship=0,
                    ))
            current_artist_id = sub_artist.id
            artist_map[sub_name] = sub_artist.id

        elif item['type'] == 'album':
            songs, ratings = _import_album(
                current_artist_id, item, tab_name, user_map, song_lookup,
            )
            total_songs += songs
            total_ratings += ratings

    db.session.commit()
    return total_songs, total_ratings


def _import_album(artist_id, album_data, tab_name, user_map, song_lookup):
    release_date = f'{album_data["year"]}-01-01' if album_data.get('year') else None
    album = Album(
        name=album_data['name'], release_date=release_date,
        album_type_id=0, submitted_by_id=0,
    )
    db.session.add(album)
    db.session.flush()

    # No genres for this import

    song_count = 0
    rating_count = 0

    for track_num, song_data in enumerate(album_data['songs'], 1):
        if song_data['is_formula_row']:
            song_lookup.setdefault('_formula_rows', []).append({
                'tab_name': tab_name, 'artist_id': artist_id,
                'song_name': song_data['name'], 'primary_tab': song_data['primary_tab'],
                'album_id': album.id, 'track_number': track_num,
                'song_data': song_data,
            })
            continue

        song = Song(name=song_data['name'], submitted_by_id=0, is_promoted=False, is_remix=False)
        db.session.add(song)
        db.session.flush()

        db.session.add(AlbumSong(album_id=album.id, song_id=song.id, track_number=track_num))
        db.session.add(ArtistSong(artist_id=artist_id, song_id=song.id, artist_is_main=True))

        for username, rating_val in song_data['ratings'].items():
            user_id = user_map.get(username)
            if user_id is not None:
                note = song_data['notes'].get(username)
                db.session.add(Rating(song_id=song.id, user_id=user_id, rating=rating_val, note=note))
                rating_count += 1

        for username, note_text in song_data['notes'].items():
            if username not in song_data['ratings']:
                user_id = user_map.get(username)
                if user_id is not None:
                    db.session.add(Rating(song_id=song.id, user_id=user_id, rating=None, note=note_text))
                    rating_count += 1

        song_lookup[(tab_name, song_data['name'])] = song.id
        song_count += 1

    return song_count, rating_count


def _link_formula_songs(wb, user_map, artist_map, song_lookup):
    formula_rows = song_lookup.pop('_formula_rows', [])
    linked = 0
    created = 0

    for frow in formula_rows:
        primary_tab = frow['primary_tab']
        song_name = frow['song_name']
        artist_id = frow['artist_id']
        base_name = re.sub(r'\s*\([^)]+\)\s*$', '', song_name).strip()
        song_id = song_lookup.get((primary_tab, song_name))

        if not song_id:
            for (tab, name), sid in song_lookup.items():
                if name == song_name:
                    song_id = sid
                    break

        if not song_id and base_name != song_name:
            song_id = song_lookup.get((primary_tab, base_name))
            if not song_id:
                for (tab, name), sid in song_lookup.items():
                    if name == base_name:
                        song_id = sid
                        break

        if not song_id and primary_tab not in wb.sheetnames:
            existing_song = Song.query.filter_by(name=song_name).first()
            if not existing_song and base_name != song_name:
                existing_song = Song.query.filter_by(name=base_name).first()
            if existing_song:
                song_id = existing_song.id

        if song_id:
            existing = ArtistSong.query.filter_by(artist_id=artist_id, song_id=song_id).first()
            if not existing:
                db.session.add(ArtistSong(artist_id=artist_id, song_id=song_id, artist_is_main=True))
                linked += 1
            existing_as = AlbumSong.query.filter_by(album_id=frow['album_id'], song_id=song_id).first()
            if not existing_as:
                db.session.add(AlbumSong(
                    album_id=frow['album_id'], song_id=song_id, track_number=frow['track_number'],
                ))
        else:
            song_data = frow['song_data']
            song = Song(name=song_name, submitted_by_id=0, is_promoted=False, is_remix=False)
            db.session.add(song)
            db.session.flush()
            db.session.add(AlbumSong(
                album_id=frow['album_id'], song_id=song.id, track_number=frow['track_number'],
            ))
            db.session.add(ArtistSong(artist_id=artist_id, song_id=song.id, artist_is_main=True))
            for username, rating_val in song_data['ratings'].items():
                user_id = user_map.get(username)
                if user_id is not None:
                    note = song_data['notes'].get(username)
                    db.session.add(Rating(song_id=song.id, user_id=user_id, rating=rating_val, note=note))
            created += 1

    db.session.commit()
    return linked, created


def _import_misc_artists(wb, user_map, artist_map, song_lookup, used_slugs):
    misc_parent = Artist.query.filter_by(name='Misc. Artists').first()
    if not misc_parent:
        print('  ERROR: "Misc. Artists" not found in DB. Run kpop import first.')
        sys.exit(1)

    misc_name = 'Misc Artists - American'
    misc_artist = Artist.query.filter_by(name=misc_name).first()
    if not misc_artist:
        slug = generate_unique_slug(misc_name, used_slugs)
        used_slugs.add(slug)
        misc_artist = Artist(
            name=misc_name, slug=slug, gender_id=2, country_id=3, submitted_by_id=0,
        )
        db.session.add(misc_artist)
        db.session.flush()
        db.session.add(ArtistArtist(
            artist_1=misc_parent.id, artist_2=misc_artist.id, relationship=0,
        ))

    album = Album(name='Misc Artists - Rock', release_date=None, album_type_id=0, submitted_by_id=0)
    db.session.add(album)
    db.session.flush()
    # Genre: Rock (id=3)
    db.session.execute(album_genres.insert().values(album_id=album.id, genre_id=3))

    total_songs = 0
    total_ratings = 0
    track_num = 0

    # Part A: "Misc Artists" tab
    if 'Misc Artists' in wb.sheetnames:
        ws = wb['Misc Artists']
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            q_cell = row[16] if len(row) > 16 else None
            if not (q_cell and q_cell.value is True and row[0].value):
                continue
            song_name = str(row[0].value).strip()
            track_num += 1
            ratings = {}
            notes = {}
            for col_idx, username in USER_COLUMNS.items():
                cell = row[col_idx]
                val, is_f, _ = _parse_rating_cell(cell)
                if val is not None:
                    ratings[username] = int(val)
                comment = _get_comment_text(cell)
                if comment:
                    notes[username] = comment

            song = Song(name=song_name, submitted_by_id=0, is_promoted=False, is_remix=False)
            db.session.add(song)
            db.session.flush()
            db.session.add(AlbumSong(album_id=album.id, song_id=song.id, track_number=track_num))
            db.session.add(ArtistSong(artist_id=misc_artist.id, song_id=song.id, artist_is_main=True))
            for username, rating_val in ratings.items():
                user_id = user_map.get(username)
                if user_id is not None:
                    note = notes.get(username)
                    db.session.add(Rating(song_id=song.id, user_id=user_id, rating=rating_val, note=note))
                    total_ratings += 1
            for username, note_text in notes.items():
                if username not in ratings:
                    user_id = user_map.get(username)
                    if user_id is not None:
                        db.session.add(Rating(song_id=song.id, user_id=user_id, rating=None, note=note_text))
                        total_ratings += 1
            total_songs += 1

    # Part B: Rated songs from skipped tabs
    for tab_name in sorted(SKIPPED_TABS):
        if tab_name not in wb.sheetnames:
            continue
        ws = wb[tab_name]
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            q_cell = row[16] if len(row) > 16 else None
            if not (q_cell and q_cell.value is True and row[0].value):
                continue
            ratings = {}
            notes = {}
            for col_idx, username in USER_COLUMNS.items():
                cell = row[col_idx]
                val, is_f, _ = _parse_rating_cell(cell)
                if val is not None:
                    ratings[username] = int(val)
                comment = _get_comment_text(cell)
                if comment:
                    notes[username] = comment
            if not ratings:
                continue
            raw_name = str(row[0].value).strip()
            song_name = f'{raw_name} ({tab_name})'
            track_num += 1
            song = Song(name=song_name, submitted_by_id=0, is_promoted=False, is_remix=False)
            db.session.add(song)
            db.session.flush()
            db.session.add(AlbumSong(album_id=album.id, song_id=song.id, track_number=track_num))
            db.session.add(ArtistSong(artist_id=misc_artist.id, song_id=song.id, artist_is_main=True))
            for username, rating_val in ratings.items():
                user_id = user_map.get(username)
                if user_id is not None:
                    note = notes.get(username)
                    db.session.add(Rating(song_id=song.id, user_id=user_id, rating=rating_val, note=note))
                    total_ratings += 1
            for username, note_text in notes.items():
                if username not in ratings:
                    user_id = user_map.get(username)
                    if user_id is not None:
                        db.session.add(Rating(song_id=song.id, user_id=user_id, rating=None, note=note_text))
                        total_ratings += 1
            total_songs += 1

    db.session.commit()
    print(f'  Misc Artists - American: {total_songs} songs, {total_ratings} ratings')
    return total_songs, total_ratings


def _import_changelog(wb):
    if 'Changelog' not in wb.sheetnames:
        return 0
    user_map = _build_user_map()
    ws = wb['Changelog']
    count = 0
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        desc_cell = row[1]
        if not desc_cell.value:
            continue
        username = str(row[0].value).strip() if row[0].value else None
        user_id = user_map.get(username) if username else None
        description = str(desc_cell.value).strip()
        date_val = None
        if row[2].value:
            if hasattr(row[2].value, 'strftime'):
                date_val = row[2].value.strftime('%Y-%m-%d')
            else:
                date_val = str(row[2].value)[:10]
        db.session.add(Changelog(
            date=date_val or '2020-01-01', user_id=user_id,
            description=description, change_type_id=3,
        ))
        count += 1
    db.session.commit()
    print(f'  {count} changelog entries imported')
    return count
