"""Import cell comment notes from 'lettuce kpop.xlsx' into existing Rating records.

Usage:
    flask import-kpop-notes

Requires: the kpop data to have been imported already (songs/ratings exist in DB).
Only imports comments from rating columns (B-L). Column A comments are skipped.
"""

import openpyxl

from app.extensions import db
from app.models.user import User
from app.models.music import Artist, Song, Rating, ArtistSong

EXCEL_PATH = 'lettuce kpop.xlsx'

# User columns B-L (0-based indices in row tuple)
USER_COLUMNS = {
    1: 'Assy',
    2: 'Stealth',
    3: 'Deren',
    4: 'Diam',
    5: 'Emily',
    6: 'Globe',
    7: 'Kanjo',
    8: 'Quiet',
    9: 'Rose',
    10: 'Sam',
    11: 'Toki',
}

# Column P (index 15) is the song flag in kpop spreadsheet
SONG_FLAG_COL = 15

META_TABS = {'Changelog', 'RULES', 'TEMPLATE', 'DIRECTORY', 'STATS', 'STATS 2.0'}


def import_kpop_notes():
    """Import cell comments from kpop spreadsheet into Rating.note fields."""
    print(f'Loading {EXCEL_PATH}...')
    wb = openpyxl.load_workbook(EXCEL_PATH)
    print(f'  {len(wb.sheetnames)} sheets found')

    user_map = {}
    for user in User.query.all():
        user_map[user.username] = user.id
    print(f'  {len(user_map)} users found in DB')

    # Build artist name → id map
    artist_map = {}
    for a in Artist.query.all():
        artist_map[a.name] = a.id

    # Build song lookup: (artist_id, song_name) → song_id
    song_lookup = {}
    for a_s in ArtistSong.query.filter_by(artist_is_main=True).all():
        song = Song.query.get(a_s.song_id)
        if song:
            song_lookup[(a_s.artist_id, song.name)] = song.id

    total_notes = 0
    total_created = 0
    total_updated = 0
    skipped_no_song = 0

    artist_tabs = set(wb.sheetnames) - META_TABS

    for tab_name in sorted(artist_tabs):
        ws = wb[tab_name]
        tab_notes = 0

        # Determine artist_id for this tab
        artist_id = artist_map.get(tab_name)

        # For Misc. Artists tab, songs have format "Song Name (Artist Name)"
        is_misc = tab_name == 'Misc. Artists'

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            a_cell = row[0]
            p_cell = row[SONG_FLAG_COL] if len(row) > SONG_FLAG_COL else None

            if not (p_cell and p_cell.value is True and a_cell.value):
                continue

            song_name = str(a_cell.value).strip()

            # Check for comments on rating columns
            for col_idx, username in USER_COLUMNS.items():
                cell = row[col_idx]
                if not cell.comment:
                    continue

                note_text = cell.comment.text.strip()
                if not note_text:
                    continue

                user_id = user_map.get(username)
                if user_id is None:
                    continue

                # Find the song
                song_id = None
                if artist_id:
                    song_id = song_lookup.get((artist_id, song_name))

                if not song_id:
                    # Try DB lookup by exact name (handles subunits, misc, etc.)
                    song = Song.query.filter_by(name=song_name).first()
                    if song:
                        song_id = song.id

                if not song_id:
                    skipped_no_song += 1
                    continue

                # Find or create Rating record
                rating = Rating.query.filter_by(song_id=song_id, user_id=user_id).first()
                if rating:
                    if not rating.note:
                        rating.note = note_text
                        total_updated += 1
                    else:
                        # Don't overwrite existing notes (e.g. from jpop import)
                        pass
                else:
                    # Create rating with note only (no score)
                    db.session.add(Rating(
                        song_id=song_id, user_id=user_id, rating=None, note=note_text,
                    ))
                    total_created += 1

                tab_notes += 1
                total_notes += 1

        if tab_notes > 0:
            db.session.commit()

    db.session.commit()
    print(f'\n=== KPOP NOTES IMPORT COMPLETE ===')
    print(f'  Total notes processed: {total_notes}')
    print(f'  Existing ratings updated with note: {total_updated}')
    print(f'  New rating rows created (note only): {total_created}')
    print(f'  Skipped (song not found in DB): {skipped_no_song}')
