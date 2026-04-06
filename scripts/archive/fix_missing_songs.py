"""Fix missing songs: import Q=False songs that sit inside album contexts.

These songs were skipped during import because their Q (song flag) column
was not checked in the spreadsheet, but they clearly belong to albums
based on their position between album headers.

Usage:
    flask fix-missing-songs
"""

import re

import openpyxl

from app.extensions import db
from app.models.user import User
from app.models.music import (
    Artist, Album, Song, Rating, ArtistSong, AlbumSong, ArtistArtist, album_genres,
)

ALBUM_YEAR_RE = re.compile(r'\((\d{4})(?:-\d{4})?\)\s*$')
IMPORTRANGE_TAB_RE = re.compile(r'""([^"!]+)!')
IFERROR_FALLBACK_RE = re.compile(r',\s*([^,)]+)\)\s*$')

USER_COLUMNS = {
    1: 'Assy', 2: 'Stealth', 3: 'Deren', 4: 'Diam', 5: 'Emily',
    6: 'Globe', 7: 'Kanjo', 8: 'Quiet', 9: 'Rose', 10: 'Sam', 11: 'Toki',
}

META_TABS_JPOP = {'Changelog', 'RULES', 'TEMPLATE', 'DIRECTORY', 'STATS', 'STATS 2.0', 'Misc. Artists'}
META_TABS_ROCK = {'Changelog', 'RULES', 'TEMPLATE', 'STATS', 'STATS 2.0', 'Misc Artists'}

# Sheets and their song-flag column index (0-based)
SHEETS = [
    ('lettuce jpop.xlsx', 16, META_TABS_JPOP),
    ('lettuce billy joel.xlsx', 16, META_TABS_ROCK),
]


def _is_formula(cell):
    return isinstance(cell.value, str) and cell.value.startswith('=')


def _parse_rating_cell(cell):
    if _is_formula(cell):
        m = IMPORTRANGE_TAB_RE.search(cell.value)
        tab = m.group(1) if m else None
        m2 = IFERROR_FALLBACK_RE.search(cell.value)
        val = None
        if m2:
            raw = m2.group(1).strip().strip('"')
            if raw and raw != '""':
                try:
                    val = float(raw)
                except ValueError:
                    pass
        return val, True, tab
    if isinstance(cell.value, (int, float)):
        v = int(cell.value)
        if 0 <= v <= 5:
            return v, False, None
    return None, False, None


def _get_comment_text(cell):
    if cell.comment:
        return cell.comment.text.strip() or None
    return None


def fix_missing_songs():
    """Find and import Q=False songs sitting inside album contexts."""
    user_map = {}
    for user in User.query.all():
        user_map[user.username] = user.id

    # Build artist lookup: name → id
    artist_map = {a.name: a.id for a in Artist.query.all()}

    # Build existing song set to avoid duplicates: (artist_id, song_name) → song_id
    existing_songs = set()
    for a_s in ArtistSong.query.all():
        song = db.session.get(Song, a_s.song_id)
        if song:
            existing_songs.add((a_s.artist_id, song.name))

    # Build album lookup: (album_name, release_date) → album_id
    album_lookup = {}
    for album in Album.query.all():
        album_lookup[(album.name, album.release_date)] = album.id

    total_added = 0
    total_ratings = 0
    total_skipped_dup = 0

    for excel_path, q_col, meta_tabs in SHEETS:
        print(f'\nProcessing {excel_path}...')
        wb = openpyxl.load_workbook(excel_path)

        for tab_name in wb.sheetnames:
            if tab_name in meta_tabs:
                continue

            ws = wb[tab_name]
            current_album_name = None
            current_album_year = None
            current_artist_id = artist_map.get(tab_name)

            # Handle tab name → artist name mapping
            if not current_artist_id:
                # Try common mappings
                name_map = {'μs': "μ's", 'NSYNC': '*NSYNC'}
                mapped = name_map.get(tab_name)
                if mapped:
                    current_artist_id = artist_map.get(mapped)

            if not current_artist_id:
                continue

            # Track subunit context (for tabs with known subunits)
            active_artist_id = current_artist_id

            tab_added = 0

            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                a_cell = row[0]
                q_cell = row[q_col] if len(row) > q_col else None
                a_val = str(a_cell.value).strip() if a_cell.value else ''

                if not a_val:
                    continue

                is_song_flag = q_cell and q_cell.value is True

                # Check if this is an album header
                m = ALBUM_YEAR_RE.search(a_val)
                if m and not is_song_flag:
                    album_name = a_val[:m.start()].strip()
                    album_year = m.group(1)
                    current_album_name = album_name
                    current_album_year = album_year

                    # Check if this row is a subunit header (no year = subunit)
                    # Actually this matched a year, so it's an album. Continue.
                    continue

                # Check if this is a known subunit header
                if not is_song_flag and not m:
                    sub_id = artist_map.get(a_val)
                    if sub_id:
                        # Check it's actually a child of the main artist
                        rel = ArtistArtist.query.filter_by(
                            artist_1=current_artist_id, artist_2=sub_id
                        ).first()
                        if rel:
                            active_artist_id = sub_id
                            current_album_name = None
                            current_album_year = None
                            continue

                if is_song_flag:
                    # Normal Q=True song — already imported, skip
                    continue

                # Q=False row inside an album context — this is a missed song
                if current_album_name is None:
                    continue

                # Find the album in DB
                release_date = f'{current_album_year}-01-01' if current_album_year else None
                album_id = album_lookup.get((current_album_name, release_date))
                if not album_id:
                    continue

                # Check for duplicate
                if (active_artist_id, a_val) in existing_songs:
                    total_skipped_dup += 1
                    continue

                # Extract ratings and notes
                ratings = {}
                notes = {}
                is_formula = False
                for col_idx, username in USER_COLUMNS.items():
                    cell = row[col_idx]
                    val, is_f, _ = _parse_rating_cell(cell)
                    if is_f:
                        is_formula = True
                    if val is not None:
                        ratings[username] = int(val)
                    comment = _get_comment_text(cell)
                    if comment:
                        notes[username] = comment

                # Skip formula rows (these are cross-artist refs, already handled)
                if is_formula:
                    continue

                # Create the song
                song = Song(name=a_val, submitted_by_id=0, is_promoted=False, is_remix=False)
                db.session.add(song)
                db.session.flush()

                # Get next track number for this album
                max_track = db.session.query(db.func.max(AlbumSong.track_number)).filter_by(
                    album_id=album_id
                ).scalar() or 0

                db.session.add(AlbumSong(
                    album_id=album_id, song_id=song.id, track_number=max_track + 1,
                ))
                db.session.add(ArtistSong(
                    artist_id=active_artist_id, song_id=song.id, artist_is_main=True,
                ))

                for username, rating_val in ratings.items():
                    user_id = user_map.get(username)
                    if user_id is not None:
                        note = notes.get(username)
                        db.session.add(Rating(
                            song_id=song.id, user_id=user_id, rating=rating_val, note=note,
                        ))
                        total_ratings += 1

                for username, note_text in notes.items():
                    if username not in ratings:
                        user_id = user_map.get(username)
                        if user_id is not None:
                            db.session.add(Rating(
                                song_id=song.id, user_id=user_id, rating=None, note=note_text,
                            ))
                            total_ratings += 1

                existing_songs.add((active_artist_id, a_val))
                total_added += 1
                tab_added += 1

            if tab_added > 0:
                db.session.commit()
                print(f'  {tab_name}: +{tab_added} songs')

    db.session.commit()
    print(f'\n=== FIX COMPLETE ===')
    print(f'  Songs added: {total_added}')
    print(f'  Ratings added: {total_ratings}')
    print(f'  Skipped (duplicate): {total_skipped_dup}')
