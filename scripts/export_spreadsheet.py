"""Export all data from the lettuce kpop spreadsheet to JSON.

Usage:
    python scripts/export_spreadsheet.py "lettuce kpop.xlsx" --output data.json

Structure rules:
- Main artist songs have s=True, album headers have s=False
- After a blank row: subunit or soloist section begins
  - First row = member/subunit name (no year in parens)
  - Subunit songs have s=True
  - Soloist songs have s=False
  - Blank row separates each member section
"""

import argparse
import json
import re
import sys

import openpyxl

NON_DATA_SHEETS = {
    'Changelog', 'RULES', 'STATS', 'STATS 2.0', 'TEMPLATE', 'DIRECTORY',
}
MISC_ARTISTS_SHEET = 'Misc. Artists'

# Tab colour → gender mapping
TAB_COLOUR_GENDER = {
    'FFFF00FF': 'female',   # Pink/Magenta
    'FF4A86E8': 'male',     # Blue
    'FF3C78D8': 'male',     # Dark Blue
    'FF00FF00': 'mixed',    # Green
}

USER_COL_START = 1   # column B (0-based)
USER_COL_END = 11    # column L (0-based, inclusive)
S_FLAG_COL = 15      # column P (0-based)


def parse_args():
    parser = argparse.ArgumentParser(description='Export spreadsheet to JSON')
    parser.add_argument('spreadsheet', help='Path to .xlsx file')
    parser.add_argument('--output', '-o', default='data.json', help='Output JSON file')
    return parser.parse_args()


def extract_users(ws):
    """Extract user list from header row of any artist sheet."""
    users = []
    header = [cell.value for cell in ws[1]]
    for i in range(USER_COL_START, USER_COL_END + 1):
        name = header[i]
        if name is None:
            continue
        is_locked = name.startswith('🔒')
        clean_name = name.replace('🔒', '').strip()
        users.append({
            'username': clean_name,
            'sort_order': len(users) + 1,
            'is_locked': is_locked,
        })
    return users


def parse_album_header(text):
    """Parse 'Album Name (Year)' → (name, year)."""
    if not text:
        return text, None
    match = re.match(r'^(.+?)\s*\((\d{4})\)\s*$', text.strip())
    if match:
        return match.group(1).strip(), match.group(2)
    return text.strip(), None


def has_year_in_parens(text):
    """Check if text contains (YYYY) pattern."""
    return bool(re.search(r'\(\d{4}\)', str(text)))


def extract_ratings(ws, row_idx, users):
    """Extract ratings for a song row. Returns dict of username→rating."""
    ratings = {}
    for i, user in enumerate(users):
        col = USER_COL_START + i + 1  # openpyxl is 1-based
        value = ws.cell(row=row_idx, column=col).value
        if value is not None:
            try:
                r = int(round(float(value)))
                if 0 <= r <= 5:
                    ratings[user['username']] = r
            except (ValueError, TypeError):
                pass
    return ratings


def extract_artist_sheet(ws, artist_name, users):
    """Extract the main artist plus any subunits/soloists from a sheet.

    Returns a list of artist entries:
    [
        {
            "name": "aespa",
            "relationship": "main",
            "parent": null,
            "albums": [...]
        },
        {
            "name": "KARINA",
            "relationship": "soloist",
            "parent": "aespa",
            "albums": [...]
        },
        {
            "name": "MISAMO",
            "relationship": "subunit",
            "parent": "TWICE",
            "albums": [...]
        },
    ]
    """
    # First pass: split the sheet into sections separated by blank rows
    sections = []
    current_section_rows = []

    for row_idx in range(2, ws.max_row + 1):
        raw_name = ws.cell(row=row_idx, column=1).value
        is_blank = raw_name is None or str(raw_name).strip() == '' or str(raw_name).strip() == 'None'

        if is_blank:
            if current_section_rows:
                sections.append(current_section_rows)
                current_section_rows = []
        else:
            s_flag = ws.cell(row=row_idx, column=S_FLAG_COL + 1).value
            current_section_rows.append((row_idx, str(raw_name).strip(), s_flag))

    if current_section_rows:
        sections.append(current_section_rows)

    if not sections:
        return []

    # First section is always the main artist
    results = []
    main_albums = _parse_section_standard(ws, sections[0], users)
    results.append({
        'name': artist_name,
        'relationship': 'main',
        'parent': None,
        'albums': [a for a in main_albums if a['songs']],
    })

    # Remaining sections are subunits or soloists
    for section in sections[1:]:
        if not section:
            continue

        # First row of section: check if it's a member header (no year)
        first_row_idx, first_name, first_s = section[0]

        if has_year_in_parens(first_name):
            # No member header — this is a continuation of albums
            # (e.g. an album that was separated by a blank row)
            albums = _parse_section_standard(ws, section, users)
            results[0]['albums'].extend([a for a in albums if a['songs']])
            continue

        # First row is a member/subunit header
        member_name = first_name
        remaining_rows = section[1:]

        if not remaining_rows:
            # Just a header with nothing under it — skip
            continue

        # Determine if subunit or soloist by checking s flags on song rows
        # Subunit songs have s=True, soloist songs have s=False
        song_s_flags = []
        for _, _, s in remaining_rows:
            if not has_year_in_parens(_):
                song_s_flags.append(s)

        # Check the non-album rows to determine type
        has_true_songs = any(s is True or s == 'True' for s in song_s_flags)
        has_false_songs = any(s is False or s == 'False' for s in song_s_flags)

        if has_true_songs:
            relationship = 'subunit'
            albums = _parse_section_standard(ws, remaining_rows, users)
        else:
            relationship = 'soloist'
            albums = _parse_section_soloist(ws, remaining_rows, users)

        albums = [a for a in albums if a['songs']]
        if albums:
            results.append({
                'name': member_name,
                'relationship': relationship,
                'parent': artist_name,
                'albums': albums,
            })

    return results


def _parse_section_standard(ws, rows, users):
    """Parse a section where s=True means song, s=False means album header."""
    albums = []
    current_album = None
    track_num = 0

    for row_idx, name_str, s_flag in rows:
        if s_flag is True or s_flag == 'True':
            track_num += 1
            ratings = extract_ratings(ws, row_idx, users)
            is_promoted = bool(ws.cell(row=row_idx, column=1).font and ws.cell(row=row_idx, column=1).font.bold)
            song = {'name': name_str, 'track_number': track_num, 'is_promoted': is_promoted, 'ratings': ratings, 'collab_artists': []}
            if current_album is not None:
                current_album['songs'].append(song)
            else:
                current_album = {'name': 'Singles', 'year': None, 'songs': [song]}
                albums.append(current_album)
        elif s_flag is False or s_flag == 'False':
            album_name, year = parse_album_header(name_str)
            current_album = {'name': album_name, 'year': year, 'songs': []}
            albums.append(current_album)
            track_num = 0

    return albums


def _parse_section_soloist(ws, rows, users):
    """Parse a soloist section where ALL rows are s=False.

    Albums have (YYYY) in the name, songs don't.
    """
    albums = []
    current_album = None
    track_num = 0

    for row_idx, name_str, s_flag in rows:
        if has_year_in_parens(name_str):
            # Album header
            album_name, year = parse_album_header(name_str)
            current_album = {'name': album_name, 'year': year, 'songs': []}
            albums.append(current_album)
            track_num = 0
        else:
            # Song
            track_num += 1
            ratings = extract_ratings(ws, row_idx, users)
            is_promoted = bool(ws.cell(row=row_idx, column=1).font and ws.cell(row=row_idx, column=1).font.bold)
            song = {'name': name_str, 'track_number': track_num, 'is_promoted': is_promoted, 'ratings': ratings, 'collab_artists': []}
            if current_album is not None:
                current_album['songs'].append(song)
            else:
                current_album = {'name': 'Singles', 'year': None, 'songs': [song]}
                albums.append(current_album)

    return albums


def extract_misc_artists(ws, users):
    """Extract songs from Misc. Artists sheet. Each song has artist in parens."""
    entries = []
    for row_idx in range(2, ws.max_row + 1):
        name = ws.cell(row=row_idx, column=1).value
        if name is None or str(name).strip() == '':
            continue

        s_flag = ws.cell(row=row_idx, column=S_FLAG_COL + 1).value
        if s_flag is not True and s_flag != 'True':
            continue

        text = str(name).strip()
        # Split on first ' (' to handle nested parens like "SHOCK (ALL(H)OURS)"
        # Format is always "Song Name (Artist Name)" with a space before the paren
        paren_idx = text.find(' (')
        if paren_idx > 0 and text.endswith(')'):
            song_name = text[:paren_idx].strip()
            artist_name = text[paren_idx + 2:-1].strip()
        else:
            song_name = text
            artist_name = 'Unknown'

        ratings = extract_ratings(ws, row_idx, users)
        entries.append({
            'song_name': song_name,
            'artist_name': artist_name,
            'ratings': ratings,
        })

    return entries


def extract_changelog(ws):
    """Extract changelog entries."""
    entries = []
    for row_idx in range(2, ws.max_row + 1):
        name = ws.cell(row=row_idx, column=1).value
        desc = ws.cell(row=row_idx, column=2).value
        date = ws.cell(row=row_idx, column=3).value

        if desc is None:
            continue

        date_str = None
        if date:
            if hasattr(date, 'strftime'):
                date_str = date.strftime('%Y-%m-%d')
            else:
                date_str = str(date)[:10]

        entries.append({
            'user': str(name).strip() if name else None,
            'description': str(desc).strip(),
            'date': date_str,
        })

    return entries


def extract_tab_genders(wb_format):
    """Extract gender from tab colours. Requires workbook opened WITHOUT data_only."""
    genders = {}
    for name in wb_format.sheetnames:
        if name in NON_DATA_SHEETS or name == MISC_ARTISTS_SHEET:
            continue
        ws = wb_format[name]
        tc = ws.sheet_properties.tabColor
        if tc and tc.rgb:
            genders[name] = TAB_COLOUR_GENDER.get(tc.rgb, 'mixed')
        else:
            genders[name] = 'mixed'
    return genders


def main():
    args = parse_args()

    print(f'Loading {args.spreadsheet}...')
    wb = openpyxl.load_workbook(args.spreadsheet, data_only=True)
    print(f'  {len(wb.sheetnames)} sheets found')

    # Load again without data_only for tab colours
    print('  Reading tab colours...')
    wb_format = openpyxl.load_workbook(args.spreadsheet)
    tab_genders = extract_tab_genders(wb_format)
    from collections import Counter
    gender_dist = Counter(tab_genders.values())
    print(f'  Gender distribution: {dict(gender_dist)}')

    artist_sheet_names = [
        name for name in wb.sheetnames
        if name not in NON_DATA_SHEETS and name != MISC_ARTISTS_SHEET
    ]

    if not artist_sheet_names:
        print('ERROR: No artist sheets found')
        sys.exit(1)

    users = extract_users(wb[artist_sheet_names[0]])
    print(f'  {len(users)} users: {[u["username"] for u in users]}')

    # Extract all artists (main + subunits + soloists)
    all_artists = []
    total_songs = 0
    total_albums = 0
    total_ratings = 0
    subunit_count = 0
    soloist_count = 0

    for sheet_name in artist_sheet_names:
        ws = wb[sheet_name]
        entries = extract_artist_sheet(ws, sheet_name, users)

        # Get gender from tab colour (main artist's tab)
        sheet_gender = tab_genders.get(sheet_name, 'mixed')

        for entry in entries:
            song_count = sum(len(a['songs']) for a in entry['albums'])
            rating_count = sum(len(s['ratings']) for a in entry['albums'] for s in a['songs'])

            if entry['albums']:
                # Subunits/soloists inherit parent's gender
                entry['gender'] = sheet_gender
                all_artists.append(entry)
                total_albums += len(entry['albums'])
                total_songs += song_count
                total_ratings += rating_count

                if entry['relationship'] == 'subunit':
                    subunit_count += 1
                elif entry['relationship'] == 'soloist':
                    soloist_count += 1

        if len([a for a in all_artists if a['relationship'] == 'main']) % 50 == 0:
            main_count = len([a for a in all_artists if a['relationship'] == 'main'])
            if main_count > 0:
                print(f'  Processed {main_count} artist sheets...')

    main_count = len([a for a in all_artists if a['relationship'] == 'main'])
    print(f'  {main_count} main artists, {subunit_count} subunits, {soloist_count} soloists')
    print(f'  {total_albums} albums, {total_songs} songs, {total_ratings} ratings')

    # Extract Misc. Artists
    misc_entries = []
    if MISC_ARTISTS_SHEET in wb.sheetnames:
        misc_entries = extract_misc_artists(wb[MISC_ARTISTS_SHEET], users)
        print(f'  {len(misc_entries)} misc artist songs')

    # Extract Changelog
    changelog = []
    if 'Changelog' in wb.sheetnames:
        changelog = extract_changelog(wb['Changelog'])
        print(f'  {len(changelog)} changelog entries')

    # Build output
    data = {
        'users': users,
        'artists': all_artists,
        'misc_artists': misc_entries,
        'changelog': changelog,
    }

    with open(args.output, 'w', encoding='utf-8') as f:
        json.dump(data, f, indent=2, ensure_ascii=False)

    print(f'\nExported to {args.output}')
    print(f'  Main artists: {main_count}')
    print(f'  Subunits: {subunit_count}')
    print(f'  Soloists: {soloist_count}')
    print(f'  Total artist entries: {len(all_artists)}')
    print(f'  Albums: {total_albums}')
    print(f'  Songs: {total_songs}')
    print(f'  Ratings: {total_ratings}')
    print(f'  Misc songs: {len(misc_entries)}')
    print(f'  Changelog: {len(changelog)}')


if __name__ == '__main__':
    main()
