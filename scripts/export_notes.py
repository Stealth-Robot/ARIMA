"""Export song-level notes (column A cell comments) from all three spreadsheets to JSON.

Usage:
    flask export-notes

Reads cell comments on song-name cells (column A) from:
  - lettuce kpop.xlsx   (song flag col P, index 15)
  - lettuce jpop.xlsx   (song flag col Q, index 16)
  - Lettuce Billy Joel.xlsx (song flag col Q, index 16)

Output: spreadsheet/notes_export.json
"""

import json
import os

import openpyxl

# ---------------------------------------------------------------------------
# Spreadsheet configs
# ---------------------------------------------------------------------------

SPREADSHEETS = [
    {
        'name': 'kpop',
        'path': os.path.join('spreadsheet', 'lettuce kpop.xlsx'),
        'song_flag_col': 15,
        'meta_tabs': {'Changelog', 'RULES', 'TEMPLATE', 'DIRECTORY', 'STATS', 'STATS 2.0'},
    },
    {
        'name': 'jpop',
        'path': os.path.join('spreadsheet', 'lettuce jpop.xlsx'),
        'song_flag_col': 16,
        'meta_tabs': {'Changelog', 'RULES', 'TEMPLATE', 'DIRECTORY', 'STATS', 'STATS 2.0', 'Misc. Artists'},
    },
    {
        'name': 'rock',
        'path': os.path.join('spreadsheet', 'Lettuce Billy Joel.xlsx'),
        'song_flag_col': 16,
        'meta_tabs': {'Changelog', 'RULES', 'TEMPLATE', 'STATS', 'STATS 2.0', 'Misc Artists'},
    },
]

OUTPUT_PATH = os.path.join('spreadsheet', 'notes_export.json')


def _scrape_notes(config):
    """Scrape column A cell comments from one spreadsheet."""
    path = config['path']
    flag_col = config['song_flag_col']
    meta_tabs = config['meta_tabs']

    print(f'  Loading {path}...')
    wb = openpyxl.load_workbook(path)
    print(f'    {len(wb.sheetnames)} sheets')

    notes = []
    artist_tabs = sorted(set(wb.sheetnames) - meta_tabs)

    for tab_name in artist_tabs:
        ws = wb[tab_name]
        tab_notes = 0

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            a_cell = row[0]
            q_cell = row[flag_col] if len(row) > flag_col else None

            if not (q_cell and q_cell.value is True and a_cell.value):
                continue

            if not a_cell.comment:
                continue

            note_text = a_cell.comment.text.strip()
            if not note_text:
                continue

            song_name = str(a_cell.value).strip()
            notes.append({
                'artist_tab': tab_name,
                'song': song_name,
                'note': note_text,
            })
            tab_notes += 1

        if tab_notes:
            print(f'    {tab_name}: {tab_notes} notes')

    print(f'    Total: {len(notes)} notes')
    return notes


def export_notes():
    """Export song-level notes from all spreadsheets to JSON."""
    print('=== EXPORT SONG NOTES ===')

    result = {}
    total = 0
    for config in SPREADSHEETS:
        print(f'\n[{config["name"].upper()}]')
        notes = _scrape_notes(config)
        result[config['name']] = notes
        total += len(notes)

    with open(OUTPUT_PATH, 'w', encoding='utf-8') as f:
        json.dump(result, f, indent=2, ensure_ascii=False)

    print(f'\n=== EXPORT COMPLETE ===')
    print(f'  Total notes: {total}')
    print(f'  Written to: {OUTPUT_PATH}')
