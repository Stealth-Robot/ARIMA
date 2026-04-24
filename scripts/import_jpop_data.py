"""Import JPOP data from 'lettuce jpop.xlsx' into the ARIMA database.

Usage:
    flask import-jpop

Requires: flask seed and the initial kpop import to have been run first.
The Excel file must be at the project root as 'lettuce jpop.xlsx'.
"""

import re
import sys

import openpyxl

from app.extensions import db
from app.models.user import User
from app.models.music import (
    Artist, Album, Song, Rating, ArtistSong, AlbumSong, ArtistArtist, album_genres,
)
from app.models.changelog import Changelog
from app.services.artist import generate_unique_slug

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

EXCEL_PATH = 'lettuce jpop.xlsx'

# User columns B–L (indices 1–11 in 0-based row tuple)
USER_COLUMNS = {
    1: 'Assy',
    2: 'Stealth',
    3: 'Deren',
    4: 'Diam',
    5: 'Emily',
    6: 'Globe',
    7: 'Kanjo',
    8: 'Quiet',
    9: 'Rose',
    10: 'Sam',
    11: 'Toki',
}

# Tabs to import as full artists (order doesn't matter)
EXPLICIT_TABS = {
    'YOASOBI', 'WEDNESDAY CAMPANELLA', 'Umamusume', 'TrySail', 'Tomori Kusunoki',
    'Tomoko Kawase', 'Tokyo Incidents', 'the peggies', 'THE ORAL CIGARETTES',
    'T.M Revolution', 'Symphogear', 'Sheena Ringo', 'Sayuri', 'Roselia',
    'Revue Starlight', 'Princess Session Orchestra', 'Pelican Fanclub',
    'Panty&Stocking', 'OLDCODEX', 'Nijigasaki Gakuen', 'Naruto', 'Nana Mizuki',
    'NANA', 'Myth & Roid', 'My Hero Academia', 'Monogatari Series', 'MIMI',
    'Lilas', 'LiSA', 'Konomi Suzuki', 'Kessoku Band', 'Kenshi Yonezu', 'Kalafina',
    'Hololive', 'Ho-kago Tea Time', 'Hikaru Utada', 'Girls Band Cry', 'GARNiDELiA',
    'FLOW', 'fripSide', 'flumpool', 'Fairy Tail', 'Eve', 'ClariS', 'BABYMETAL',
    'Ayumi Hamasaki', 'Bleach', 'Ayahi Takagaki', 'ASIAN KUNG-FU GENERATION',
    'Aqours', 'Ami Suzuki', 'Aimer', 'Ado',
    'μs',  # sheet name (no apostrophe)
}

# Tabs that are metadata, not artist data
META_TABS = {'Changelog', 'RULES', 'TEMPLATE', 'DIRECTORY', 'STATS', 'STATS 2.0', 'Misc. Artists'}

# Display name overrides (sheet name → artist name in DB)
TAB_NAME_MAP = {
    'μs': "μ's",
}

# Gender: 0=Female, 1=Male, 2=Mixed
GENDER_MAP = {
    'Ado': 0, 'Aimer': 0, 'Ami Suzuki': 0, 'Aqours': 0, 'Ayahi Takagaki': 0,
    'Ayumi Hamasaki': 0, 'ClariS': 0, 'fripSide': 0, 'Girls Band Cry': 0,
    'Ho-kago Tea Time': 0, 'Hikaru Utada': 0, 'Kalafina': 0, 'Kessoku Band': 0,
    'Konomi Suzuki': 0, 'Lilas': 0, 'LiSA': 0, 'MIMI': 0, 'Nana Mizuki': 0,
    'Nijigasaki Gakuen': 0, 'Princess Session Orchestra': 0, 'Roselia': 0,
    'Sayuri': 0, 'Sheena Ringo': 0, 'the peggies': 0, 'Tomoko Kawase': 0,
    'Tomori Kusunoki': 0, 'TrySail': 0, "μ's": 0,
    'ASIAN KUNG-FU GENERATION': 1, 'Eve': 1, 'FLOW': 1, 'flumpool': 1,
    'Kenshi Yonezu': 1, 'OLDCODEX': 1, 'Pelican Fanclub': 1,
    'T.M Revolution': 1, 'THE ORAL CIGARETTES': 1,
    'BABYMETAL': 2, 'Bleach': 2, 'Fairy Tail': 2, 'GARNiDELiA': 2,
    'Hololive': 2, 'Monogatari Series': 2, 'My Hero Academia': 2,
    'Myth & Roid': 2, 'NANA': 2, 'Naruto': 2, 'Panty&Stocking': 2,
    'Revue Starlight': 2, 'Symphogear': 2, 'Tokyo Incidents': 2,
    'Umamusume': 2, 'WEDNESDAY CAMPANELLA': 2, 'YOASOBI': 2,
}

# Album row detection: "Name (YYYY)" or "Name (YYYY-YYYY)"
ALBUM_YEAR_RE = re.compile(r'\((\d{4})(?:-\d{4})?\)\s*$')

# IMPORTRANGE formula pattern to extract referenced tab name
IMPORTRANGE_TAB_RE = re.compile(r'""([^"!]+)!')

# IFERROR fallback value extraction: last arg before closing paren
IFERROR_FALLBACK_RE = re.compile(r',\s*([^,)]+)\)\s*$')

# Hardcoded subunit names per tab (auto-detection is too fragile)
KNOWN_SUBUNITS = {
    'Hololive': {
        'hololive 5th Generation', 'Blue Journey', 'Suisei Hoshimachi',
        'Marine Houshou', 'Polka Omaru', 'Shiranui Constructions',
        'FUWAMOCO', 'COVER SONGS', 'Sakura Miko',
    },
}


# ---------------------------------------------------------------------------
# Parsing helpers
# ---------------------------------------------------------------------------

def _parse_album_year(text):
    """Extract year from album row text. Returns (name, year_str) or None."""
    m = ALBUM_YEAR_RE.search(text)
    if not m:
        return None
    year = m.group(1)
    name = text[:m.start()].strip()
    return name, year


def _is_formula(cell):
    """Check if a cell contains a formula string."""
    return isinstance(cell.value, str) and cell.value.startswith('=')


def _extract_formula_tab(cell):
    """Extract the referenced tab name from an IMPORTRANGE formula."""
    if not _is_formula(cell):
        return None
    m = IMPORTRANGE_TAB_RE.search(cell.value)
    return m.group(1) if m else None


def _extract_formula_fallback(cell):
    """Extract the numeric fallback from an IFERROR formula. Returns float or None."""
    if not _is_formula(cell):
        return None
    m = IFERROR_FALLBACK_RE.search(cell.value)
    if not m:
        return None
    raw = m.group(1).strip().strip('"')
    if raw == '' or raw == '""':
        return None
    try:
        return float(raw)
    except ValueError:
        return None


def _parse_rating_cell(cell):
    """Return (rating_value_or_None, is_formula, referenced_tab_or_None)."""
    if _is_formula(cell):
        tab = _extract_formula_tab(cell)
        val = _extract_formula_fallback(cell)
        return val, True, tab
    if isinstance(cell.value, (int, float)):
        v = int(cell.value)
        if 0 <= v <= 5:
            return v, False, None
    return None, False, None


def _get_comment_text(cell):
    """Return cell comment text or None."""
    if cell.comment:
        return cell.comment.text.strip() or None
    return None


def _extract_song_data(row):
    """Extract ratings, notes, formula info from a song row. Returns dict."""
    ratings = {}
    notes = {}
    is_formula_row = False
    primary_tab = None

    for col_idx, username in USER_COLUMNS.items():
        cell = row[col_idx]
        val, is_f, tab = _parse_rating_cell(cell)
        if is_f:
            is_formula_row = True
            if tab:
                primary_tab = tab
        if val is not None:
            ratings[username] = int(val)
        comment = _get_comment_text(cell)
        if comment:
            notes[username] = comment

    return {
        'ratings': ratings, 'notes': notes,
        'is_formula_row': is_formula_row, 'primary_tab': primary_tab,
    }


def _row_has_ratings(row):
    """Check if a row has any valid numeric rating in user columns B-L."""
    for col_idx in USER_COLUMNS:
        cell = row[col_idx]
        if isinstance(cell.value, (int, float)) and 0 <= cell.value <= 5:
            return True
        if _is_formula(cell):
            return True
    return False


def _parse_tab(ws, known_subunits=None):
    """Parse a single artist worksheet into structured data.

    Args:
        ws: openpyxl worksheet
        known_subunits: optional set of known subunit names for this tab

    Returns a list of items, each either:
      {'type': 'subunit', 'name': str}
      {'type': 'album', 'name': str, 'year': str|None, 'songs': [...]}
    where each song is:
      {'name': str, 'ratings': {username: int}, 'notes': {username: str},
       'is_formula_row': bool, 'primary_tab': str|None}
    """
    subunits = known_subunits or set()
    in_subunit_tab = bool(subunits)
    items = []
    current_album = None

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        a_cell = row[0]
        q_cell = row[16] if len(row) > 16 else None
        a_val = str(a_cell.value).strip() if a_cell.value else ''
        is_song = q_cell and q_cell.value is True

        if not a_val:
            continue  # spacer

        if is_song:
            # Definitely a song
            song_data = _extract_song_data(row)
            song_data['name'] = a_val
            if current_album:
                current_album['songs'].append(song_data)
            else:
                current_album = {
                    'type': 'album', 'name': 'Unknown Album', 'year': None,
                    'songs': [song_data], '_implicit': True,
                }
                items.append(current_album)
            continue

        # Q=False row with text
        parsed = _parse_album_year(a_val)
        if parsed:
            name, year = parsed
            current_album = {
                'type': 'album', 'name': name, 'year': year, 'songs': [],
            }
            items.append(current_album)
            continue

        # No year pattern — check against known subunits first
        if a_val in subunits:
            items.append({'type': 'subunit', 'name': a_val})
            current_album = None
            continue

        # Has ratings → treat as a misclassified song (Q was wrong)
        if _row_has_ratings(row):
            song_data = _extract_song_data(row)
            song_data['name'] = a_val
            if current_album:
                current_album['songs'].append(song_data)
            else:
                current_album = {
                    'type': 'album', 'name': 'Unknown Album', 'year': None,
                    'songs': [song_data], '_implicit': True,
                }
                items.append(current_album)
            continue

        # No ratings, no year, not a known subunit.
        # In tabs with subunits, treat as a song (e.g. unreleased singles under a subunit).
        # In other tabs, skip (likely a TODO note).
        if in_subunit_tab:
            song_data = _extract_song_data(row)
            song_data['name'] = a_val
            if current_album:
                current_album['songs'].append(song_data)
            else:
                current_album = {
                    'type': 'album', 'name': 'Unknown Album', 'year': None,
                    'songs': [song_data], '_implicit': True,
                }
                items.append(current_album)

    return items


# ---------------------------------------------------------------------------
# Import logic
# ---------------------------------------------------------------------------

def import_jpop():
    """Main entry point — import all JPOP data from the Excel file."""
    print(f'Loading {EXCEL_PATH} (formulas mode)...')
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=False)
    print(f'  {len(wb.sheetnames)} sheets found')

    # Build user map from existing DB users
    user_map = _build_user_map()
    used_slugs = {a.slug for a in Artist.query.all() if a.slug}

    # Track artist_name → artist_id for multi-artist linking
    artist_map = {}
    # Track (tab_name, song_name) → song_id for formula-row linking
    song_lookup = {}

    total_songs = 0
    total_ratings = 0

    # --- Phase 1: Import explicit artist tabs ---
    for tab_name in sorted(EXPLICIT_TABS):
        if tab_name not in wb.sheetnames:
            print(f'  WARNING: tab "{tab_name}" not found in workbook, skipping')
            continue

        artist_name = TAB_NAME_MAP.get(tab_name, tab_name)
        ws = wb[tab_name]
        items = _parse_tab(ws, known_subunits=KNOWN_SUBUNITS.get(tab_name))

        songs, ratings = _import_artist_tab(
            artist_name, tab_name, items, user_map, artist_map, song_lookup, used_slugs,
        )
        total_songs += songs
        total_ratings += ratings
        print(f'  {tab_name}: {songs} songs, {ratings} ratings')

    # --- Phase 2: Link multi-artist (formula) songs ---
    linked, formula_created = _link_formula_songs(wb, user_map, artist_map, song_lookup)
    total_songs += formula_created
    print(f'  Multi-artist links: {linked}, formula songs created as fallback: {formula_created}')

    # --- Phase 3: Misc. Artists ---
    misc_songs, misc_ratings = _import_misc_artists(wb, user_map, artist_map, song_lookup, used_slugs)
    total_songs += misc_songs
    total_ratings += misc_ratings

    # --- Phase 4: Changelog ---
    cl_count = _import_changelog(wb)

    print(f'\n=== JPOP IMPORT COMPLETE ===')
    print(f'  Artists: {len(artist_map)}')
    print(f'  Songs: {total_songs}')
    print(f'  Ratings: {total_ratings}')
    print(f'  Multi-artist links: {linked} (+ {formula_created} created as fallback)')
    print(f'  Changelog: {cl_count}')


def _build_user_map():
    """Build {username: user_id} from existing DB users."""
    user_map = {}
    for user in User.query.all():
        user_map[user.username] = user.id
    print(f'  {len(user_map)} users found in DB')
    return user_map


def _import_artist_tab(artist_name, tab_name, items, user_map, artist_map, song_lookup, used_slugs):
    """Import a single artist tab. Creates artist, subunits, albums, songs, ratings."""
    gender_id = GENDER_MAP.get(artist_name, 2)

    # Create or find main artist
    main_artist = Artist.query.filter_by(name=artist_name).first()
    if not main_artist:
        slug = generate_unique_slug(artist_name, used_slugs)
        used_slugs.add(slug)
        main_artist = Artist(
            name=artist_name, slug=slug, gender_id=gender_id,
            country_id=1, submitted_by_id=0,
        )
        db.session.add(main_artist)
        db.session.flush()

    artist_map[tab_name] = main_artist.id
    artist_map[artist_name] = main_artist.id

    total_songs = 0
    total_ratings = 0
    current_artist_id = main_artist.id

    for item in items:
        if item['type'] == 'subunit':
            # Create subunit artist
            sub_name = item['name']
            sub_artist = Artist.query.filter_by(name=sub_name).first()
            if not sub_artist:
                slug = generate_unique_slug(sub_name, used_slugs)
                used_slugs.add(slug)
                sub_artist = Artist(
                    name=sub_name, slug=slug, gender_id=gender_id,
                    country_id=1, submitted_by_id=0,
                )
                db.session.add(sub_artist)
                db.session.flush()
                # Create ArtistArtist link
                existing_rel = ArtistArtist.query.filter_by(
                    artist_1=main_artist.id, artist_2=sub_artist.id,
                ).first()
                if not existing_rel:
                    db.session.add(ArtistArtist(
                        artist_1=main_artist.id, artist_2=sub_artist.id, relationship=0,
                    ))
            current_artist_id = sub_artist.id
            artist_map[sub_name] = sub_artist.id

        elif item['type'] == 'album':
            songs, ratings = _import_album(
                current_artist_id, item, tab_name, user_map, song_lookup,
            )
            total_songs += songs
            total_ratings += ratings

    db.session.commit()
    return total_songs, total_ratings


def _import_album(artist_id, album_data, tab_name, user_map, song_lookup):
    """Create album + songs + ratings. Returns (song_count, rating_count)."""
    release_date = f'{album_data["year"]}-01-01' if album_data.get('year') else None
    album = Album(
        name=album_data['name'], release_date=release_date,
        album_type_id=0, submitted_by_id=0,
    )
    db.session.add(album)
    db.session.flush()

    # Genre: Jpop (id=1)
    db.session.execute(album_genres.insert().values(album_id=album.id, genre_id=1))

    song_count = 0
    rating_count = 0

    for track_num, song_data in enumerate(album_data['songs'], 1):
        if song_data['is_formula_row']:
            # This song belongs to another artist's tab — try to link in Phase 2.
            # Store full song_data so we can create it as fallback if linking fails.
            song_lookup.setdefault('_formula_rows', []).append({
                'tab_name': tab_name,
                'artist_id': artist_id,
                'song_name': song_data['name'],
                'primary_tab': song_data['primary_tab'],
                'album_id': album.id,
                'track_number': track_num,
                'song_data': song_data,
            })
            continue

        song = Song(
            name=song_data['name'], submitted_by_id=0,
            is_promoted=False, is_remix=False,
        )
        db.session.add(song)
        db.session.flush()

        db.session.add(AlbumSong(
            album_id=album.id, song_id=song.id, track_number=track_num,
        ))
        db.session.add(ArtistSong(
            artist_id=artist_id, song_id=song.id, artist_is_main=True,
        ))

        # Ratings + notes
        for username, rating_val in song_data['ratings'].items():
            user_id = user_map.get(username)
            if user_id is not None:
                note = song_data['notes'].get(username)
                db.session.add(Rating(
                    song_id=song.id, user_id=user_id, rating=rating_val, note=note,
                ))
                rating_count += 1

        # Notes without ratings (user left a comment but no score)
        for username, note_text in song_data['notes'].items():
            if username not in song_data['ratings']:
                user_id = user_map.get(username)
                if user_id is not None:
                    db.session.add(Rating(
                        song_id=song.id, user_id=user_id, rating=None, note=note_text,
                    ))
                    rating_count += 1

        # Register for cross-tab lookup
        song_lookup[(tab_name, song_data['name'])] = song.id
        song_count += 1

    return song_count, rating_count


def _link_formula_songs(wb, user_map, artist_map, song_lookup):
    """Phase 2: Link formula rows to their primary songs in other tabs.

    If a primary song can't be found (e.g. referenced tab was skipped),
    create the song as a regular entry under the current artist instead.
    Returns (linked_count, created_count).
    """
    formula_rows = song_lookup.pop('_formula_rows', [])
    linked = 0
    created = 0

    for frow in formula_rows:
        primary_tab = frow['primary_tab']
        song_name = frow['song_name']
        artist_id = frow['artist_id']

        # Find the primary song by (primary_tab, song_name)
        base_name = re.sub(r'\s*\([^)]+\)\s*$', '', song_name).strip()
        song_id = song_lookup.get((primary_tab, song_name))

        if not song_id:
            # Try fuzzy: match by song name across all keys
            for (tab, name), sid in song_lookup.items():
                if name == song_name:
                    song_id = sid
                    break

        if not song_id:
            # Try matching without parenthetical suffix:
            # "NEVER-END TALE (Konomi Suzuki)" → look for "NEVER-END TALE"
            if base_name != song_name:
                song_id = song_lookup.get((primary_tab, base_name))
                if not song_id:
                    for (tab, name), sid in song_lookup.items():
                        if name == base_name:
                            song_id = sid
                            break

        if not song_id and primary_tab not in wb.sheetnames:
            # Cross-spreadsheet ref (e.g. BoA from kpop sheet) — try DB lookup
            existing_song = Song.query.filter_by(name=song_name).first()
            if not existing_song and base_name != song_name:
                existing_song = Song.query.filter_by(name=base_name).first()
            if existing_song:
                song_id = existing_song.id

        if song_id:
            # Link this artist to the existing song
            existing = ArtistSong.query.filter_by(
                artist_id=artist_id, song_id=song_id,
            ).first()
            if not existing:
                db.session.add(ArtistSong(
                    artist_id=artist_id, song_id=song_id, artist_is_main=True,
                ))
                linked += 1

            # Also add AlbumSong link so it appears in this artist's album
            existing_as = AlbumSong.query.filter_by(
                album_id=frow['album_id'], song_id=song_id,
            ).first()
            if not existing_as:
                db.session.add(AlbumSong(
                    album_id=frow['album_id'], song_id=song_id,
                    track_number=frow['track_number'],
                ))
        else:
            # Primary song not found — create as regular song under this artist
            song_data = frow['song_data']
            song = Song(
                name=song_name, submitted_by_id=0,
                is_promoted=False, is_remix=False,
            )
            db.session.add(song)
            db.session.flush()

            db.session.add(AlbumSong(
                album_id=frow['album_id'], song_id=song.id,
                track_number=frow['track_number'],
            ))
            db.session.add(ArtistSong(
                artist_id=artist_id, song_id=song.id, artist_is_main=True,
            ))

            for username, rating_val in song_data['ratings'].items():
                user_id = user_map.get(username)
                if user_id is not None:
                    note = song_data['notes'].get(username)
                    db.session.add(Rating(
                        song_id=song.id, user_id=user_id,
                        rating=rating_val, note=note,
                    ))

            created += 1

    db.session.commit()
    return linked, created


def _import_misc_artists(wb, user_map, artist_map, song_lookup, used_slugs):
    """Phase 3: Import Misc. Artists tab + rated songs from skipped tabs."""
    # Find existing "Misc. Artists" parent
    misc_parent = Artist.query.filter_by(name='Misc. Artists').first()
    if not misc_parent:
        print('  ERROR: "Misc. Artists" not found in DB. Run kpop import first.')
        sys.exit(1)

    # Create "Misc. Artists - Japanese" subunit
    jp_misc_name = 'Misc. Artists - Japanese'
    jp_misc = Artist.query.filter_by(name=jp_misc_name).first()
    if not jp_misc:
        slug = generate_unique_slug(jp_misc_name, used_slugs)
        used_slugs.add(slug)
        jp_misc = Artist(
            name=jp_misc_name, slug=slug, gender_id=2,
            country_id=1, submitted_by_id=0,
        )
        db.session.add(jp_misc)
        db.session.flush()
        db.session.add(ArtistArtist(
            artist_1=misc_parent.id, artist_2=jp_misc.id, relationship=0,
        ))

    # Create "Misc. Artists - JPOP" album
    jp_album = Album(
        name='Misc. Artists - JPOP', release_date=None,
        album_type_id=0, submitted_by_id=0,
    )
    db.session.add(jp_album)
    db.session.flush()
    db.session.execute(album_genres.insert().values(album_id=jp_album.id, genre_id=1))

    total_songs = 0
    total_ratings = 0
    track_num = 0

    # --- Part A: Songs from the "Misc. Artists" spreadsheet tab ---
    if 'Misc. Artists' in wb.sheetnames:
        ws = wb['Misc. Artists']
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            a_cell = row[0]
            q_cell = row[16] if len(row) > 16 else None

            if not (q_cell and q_cell.value is True and a_cell.value):
                continue

            song_name = str(a_cell.value).strip()
            track_num += 1

            ratings = {}
            notes = {}
            for col_idx, username in USER_COLUMNS.items():
                cell = row[col_idx]
                val, is_f, _ = _parse_rating_cell(cell)
                if val is not None:
                    ratings[username] = int(val)
                comment = _get_comment_text(cell)
                if comment:
                    notes[username] = comment

            song = Song(name=song_name, submitted_by_id=0, is_promoted=False, is_remix=False)
            db.session.add(song)
            db.session.flush()

            db.session.add(AlbumSong(album_id=jp_album.id, song_id=song.id, track_number=track_num))
            db.session.add(ArtistSong(artist_id=jp_misc.id, song_id=song.id, artist_is_main=True))

            for username, rating_val in ratings.items():
                user_id = user_map.get(username)
                if user_id is not None:
                    note = notes.get(username)
                    db.session.add(Rating(
                        song_id=song.id, user_id=user_id, rating=rating_val, note=note,
                    ))
                    total_ratings += 1

            for username, note_text in notes.items():
                if username not in ratings:
                    user_id = user_map.get(username)
                    if user_id is not None:
                        db.session.add(Rating(
                            song_id=song.id, user_id=user_id, rating=None, note=note_text,
                        ))
                        total_ratings += 1

            total_songs += 1

    # --- Part B: Rated songs from skipped tabs ---
    all_tabs = set(wb.sheetnames)
    skipped_tabs = sorted(all_tabs - EXPLICIT_TABS - META_TABS)

    for tab_name in skipped_tabs:
        ws = wb[tab_name]
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            a_cell = row[0]
            q_cell = row[16] if len(row) > 16 else None

            if not (q_cell and q_cell.value is True and a_cell.value):
                continue

            # Check if song has any rating
            ratings = {}
            notes = {}
            for col_idx, username in USER_COLUMNS.items():
                cell = row[col_idx]
                val, is_f, _ = _parse_rating_cell(cell)
                if val is not None:
                    ratings[username] = int(val)
                comment = _get_comment_text(cell)
                if comment:
                    notes[username] = comment

            if not ratings:
                continue  # Only import rated songs from skipped tabs

            raw_name = str(a_cell.value).strip()
            song_name = f'{raw_name} ({tab_name})'
            track_num += 1

            song = Song(name=song_name, submitted_by_id=0, is_promoted=False, is_remix=False)
            db.session.add(song)
            db.session.flush()

            db.session.add(AlbumSong(album_id=jp_album.id, song_id=song.id, track_number=track_num))
            db.session.add(ArtistSong(artist_id=jp_misc.id, song_id=song.id, artist_is_main=True))

            for username, rating_val in ratings.items():
                user_id = user_map.get(username)
                if user_id is not None:
                    note = notes.get(username)
                    db.session.add(Rating(
                        song_id=song.id, user_id=user_id, rating=rating_val, note=note,
                    ))
                    total_ratings += 1

            for username, note_text in notes.items():
                if username not in ratings:
                    user_id = user_map.get(username)
                    if user_id is not None:
                        db.session.add(Rating(
                            song_id=song.id, user_id=user_id, rating=None, note=note_text,
                        ))
                        total_ratings += 1

            total_songs += 1

    db.session.commit()
    print(f'  Misc. Artists - Japanese: {total_songs} songs, {total_ratings} ratings')
    return total_songs, total_ratings


def _import_changelog(wb):
    """Phase 4: Import changelog entries."""
    if 'Changelog' not in wb.sheetnames:
        return 0

    user_map = _build_user_map()
    ws = wb['Changelog']
    count = 0

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        name_cell = row[0]  # A
        desc_cell = row[1]  # B
        date_cell = row[2]  # C

        if not desc_cell.value:
            continue

        username = str(name_cell.value).strip() if name_cell.value else None
        user_id = user_map.get(username) if username else None
        description = str(desc_cell.value).strip()

        date_val = None
        if date_cell.value:
            if hasattr(date_cell.value, 'strftime'):
                date_val = date_cell.value.strftime('%Y-%m-%d')
            else:
                date_val = str(date_cell.value)[:10]

        db.session.add(Changelog(
            date=date_val or '2020-01-01',
            user_id=user_id,
            description=description,
            change_type_id=3,  # Legacy
        ))
        count += 1

    db.session.commit()
    print(f'  {count} changelog entries imported')
    return count
